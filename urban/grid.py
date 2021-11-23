import numpy as np
from openpyxl import Workbook
import math
import os
import sys
import pandas as pd
from tensorflow.keras.models import load_model
import matplotlib.pyplot as plt
from PIL import Image
x_min = 324600      # minimum x value
y_min = 4150520     # minimum y value


z_max = 288.43

# for grayscale
under_ground = 255.0
in_terrain = 200.0
in_building = 200.0
in_sky = 0.0


def load_npy():
    dim_1 = np.load('./maps/Building.npy')
    dim_2 = np.load('./maps/Terrain.npy')

    return dim_1, dim_2


dim_1, dim_2 = load_npy()


def discriminant(a, b, c):
    return ((-b + math.sqrt(b ** 2 - 4 * a * c)) / (2 * a), (-b - math.sqrt(b ** 2 - 4 * a * c)) / (2 * a))


def line_equation(a, d, gradient, z):
    return (a, a * gradient + d, z)


def sol(x1, y1, z1, gradient, width):
    d = y1 - gradient * x1
    a = gradient ** 2 + 1
    b = -2 * x1 + 2 * gradient * d - 2 * y1 * gradient
    c = x1 ** 2 + d ** 2 - 2 * y1 * d + y1 ** 2 - (width/2)**2
    return (
        line_equation(discriminant(a, b, c)[1], d, gradient, z1),
        line_equation(discriminant(a, b, c)[0], d, gradient, z1))


def gen_image(Tx_coord, Rx_coord, name, img_width, img_height, distance, ver, chai):
    Tx_coord[0] = Tx_coord[0] - x_min
    Tx_coord[1] = Tx_coord[1] - y_min
    Rx_coord[0] = Rx_coord[0] - x_min
    Rx_coord[1] = Rx_coord[1] - y_min
    # print(Tx_coord[2], ' -> ', end='')
    # Tx_coord[2] += dim_1[Tx_coord[0]][Tx_coord[1]]
    # print(Tx_coord[2], ' -> ', end='')
    # Rx_coord[2] = dim_1[Rx_coord[1]][Rx_coord[0]]
    if distance % 10 != 0:
        alpha = (Tx_coord[0] - Rx_coord[0]) / (distance - 3)
        beta = (Tx_coord[1] - Rx_coord[1]) / (distance - 3)
        Tx_coord[0] += alpha
        Tx_coord[1] += alpha
        Rx_coord[0] -= beta
        Rx_coord[1] -= beta
    # Tx_coord[2] += dim_1[Tx_coord[0]][Tx_coord[1]]
    # Rx_coord[2] = dim_1[Rx_coord[0]][Rx_coord[1]]
    if Tx_coord[1] == Rx_coord[1]:
        XYs = [[]]
        CentreDots = np.linspace(Tx_coord, Rx_coord, distance)
        for i in range(distance):
            XYs.append(getDots_Of_Orthogonal_Line(CentreDots[i], None, img_width, distance))
        if ver == 1:
            array_to_npy(XYs, Tx_coord, Rx_coord, name, img_width, img_height, distance)
        elif ver == 2:
            array_to_npy2(XYs, Tx_coord, Rx_coord, name, img_width, img_height, distance, chai)
        return
    gradient = (Tx_coord[0] - Rx_coord[0]) / (Tx_coord[1] - Rx_coord[1])
    gradient *= (-1)
    CentreDots = getDottedLines(Tx_coord, Rx_coord, distance)  # CentreDots,

    # Tx와 Rx를 잇는 직선위의 점들(distance로 uniform하게 나눠짐)
    # getDottedLines의 인자의 dim=2였으므로, 2차원 배열 반환
    XYs = [[]]
    for i in range(len(CentreDots)):  # 직선위의 점을 지나고, TxRx 직선을 수직하는 선분의 양 끝점들을 받아오기
        XYs.append(getDots_Of_Orthogonal_Line(CentreDots[i], gradient, img_width, distance))



    if ver == 1:
        array_to_npy(XYs, Tx_coord, Rx_coord, name, img_width, img_height, distance)
    elif ver == 2:
        array_to_npy2(XYs, Tx_coord, Rx_coord, name, img_width, img_height, distance, chai)


def getDottedLines(Tx, Rx, NumOfSlice):
    return np.linspace(Tx, Rx, NumOfSlice)


def getDots_Of_Orthogonal_Line(CenterPoints, gradient, img_width, distance):
    if gradient is None:
        twoDots = []
        twoDots.append((CenterPoints[0], CenterPoints[1] - (img_width / 2), CenterPoints[2]))
        twoDots.append((CenterPoints[0], CenterPoints[1] + (img_width / 2), CenterPoints[2]))
        return getDottedLines(twoDots[0], twoDots[1], img_width)
    twoDots = sol(CenterPoints[0], CenterPoints[1], CenterPoints[2], gradient, img_width)
    return getDottedLines(twoDots[0], twoDots[1], img_width)


def array_to_npy(XYs, Tx, Rx, name, img_width, img_height, distance):
    cube = np.arange(img_width * img_height * distance).reshape((img_width, img_height, distance))
    f = open("temp.txt", "w")
    for i in range(distance):
        for j in range(img_width):
            for k in range(img_height):
                cube[j][k][i] = 0

    for i in range(1, distance + 1):
        for j in range(img_width):
            for k in range(img_height):
                try:
                    x = int(np.rint(XYs[i][j][0]))
                    y = int(np.rint(XYs[i][j][1]))
                    z = int(np.rint(XYs[i][j][2] - (img_height / 2) + k))

                    if z < 0:
                        cube[k][j][i - 1] = under_ground
                    elif z < int(dim_1[y][x]):
                        cube[k][j][i - 1] = in_terrain
                    elif z < int(dim_2[y][x]):
                        cube[k][j][i - 1] = in_building
                    else:
                        cube[k][j][i - 1] = in_sky
                    if cube[k][j][i - 1] > 255:
                        cube[k][j][i - 1] = 255
                    # cube[k][j][i-1] /= 255

                    # print(str(x + x_min) + '\t' + str(y + y_min))
                except:
                    print(i, j, k)
                    print(cube.shape)
                    print(img_width, img_height, distance)
                    exit(1)
    cube = flip_or_not(cube, Tx[0] - Rx[0], Tx[1] - Rx[1], img_width, img_height, distance)
    np.save(name + '.npy', cube)

def array_to_npy2(XYs, Tx, Rx, name, img_width, img_height, distance, chai):
    cube_org = np.arange(img_width * img_height * distance).reshape((img_width, img_height, distance))
    # d = np.sqrt(np.power(Tx[0] - Rx[0], 2) + np.power(Tx[1] - Rx[1], 2)) / distance
    d = int((180 - chai)/10)
    # d *= 1.35
    for i in range(distance):
        for j in range(img_width):
            for k in range(img_height):
                cube_org[j][k][i] = 0

    for i in range(1, distance + 1):
        for j in range(img_width):
            for k in range(img_height):
                try:
                    x = int(np.rint(XYs[i][j][0]))
                    y = int(np.rint(XYs[i][j][1]))
                    z = int(np.rint(XYs[i][j][2] - (img_height / 2) + k))
                    if z < 0:
                        cube_org[k][j][i - 1] = under_ground + d
                    elif z < int(dim_1[y][x]):
                        cube_org[k][j][i - 1] = in_terrain + d
                    elif z < int(dim_2[y][x]):
                        cube_org[k][j][i - 1] = in_building + d
                    else:
                        cube_org[k][j][i - 1] = in_sky + d
                    if cube_org[k][j][i - 1] > 255:
                        cube_org[k][j][i - 1] = 255
                except:
                    print(i, j, k)
                    exit(1)
    cube_org = flip_or_not(cube_org, Tx[0] - Rx[0], Tx[1] - Rx[1], img_width, img_height, distance)
    origin_size = 40
    change_size = 20
    ratio = int(origin_size / change_size)
    cube = np.arange(change_size * change_size * change_size).reshape(change_size, change_size, change_size)
    for dist in range(change_size):
        for i in range(change_size):
            for j in range(change_size):
                cube[i][j][dist] = int((cube_org[i * ratio][j * ratio][dist * ratio + 1] +
                                        cube_org[i * ratio + 1][j * ratio][dist * ratio + 1] +
                                        cube_org[i * ratio][j * ratio + 1][dist * ratio + 1] +
                                        cube_org[i * ratio + 1][j * ratio + 1][dist * ratio + 1]) / 4)

    # np.save(lamsdir + new_target + name, cube)
    np.save(name + '.npy', cube)


def flip_or_not(cube, a, b, img_width, img_height, distance):
    case = 0
    if a > 0:
        if b > 0:
            case = 1
    elif a < 0:
        if b >= 0:
            case = 1
    else:
        if b > 0:
            case = 1
    if case == 0:
        return cube
    if case == 1:
        f_cube = np.arange(img_width * img_height * distance).reshape((img_width, img_height, distance))
        for i in range(img_height):
            for j in range(img_width):
                for k in range(distance):
                    f_cube[i][j][k] = 0

        for i in range(img_height):
            for j in range(img_width):
                for k in range(distance):
                    try:
                        f_cube[i][j][k] = cube[i][j][img_width - k - 1]
                    except:
                        print(cube.shape, distance, img_width, img_height)
        return f_cube


def angle_diff(Tx, Rx, azi, dtilt):
    ln1 = Rx[0] - Tx[0]
    ln2 = Rx[1] - Tx[1]
    ln3 = Tx[2] - Rx[2]
    dist2d = math.sqrt(math.pow(ln1, 2) + math.pow(ln2, 2))

    theta = math.atan2(ln2, ln1)
    theta = theta * 180 / math.pi

    theta2 = math.atan2(ln3, dist2d)
    theta2 = theta2 * 180 / math.pi
    case = 0
    if ln2 > 0:
        if ln1 > 0:
            theta = 90 - theta
            case = 1
        else:
            theta = 450 - theta
            case = 2
    else:
        theta = 90 + theta * (-1)
        case = 3
    azi_diff = abs(theta - azi)
    tilt_diff = abs(theta2 - dtilt)

    alpha = ln3 / math.tan(dtilt)
    gamma = math.sqrt(math.pow(alpha, 2) + math.pow(ln3, 2))
    beta = math.sqrt(math.pow(alpha*math.sin(azi_diff), 2) + math.pow(dist2d - alpha * math.cos(azi_diff), 2))
    dist3d = math.sqrt(math.pow(ln1, 2) + math.pow(ln2, 2) + math.pow(ln3, 2))
    f_diff = math.acos((math.pow(dist3d, 2) + math.pow(gamma, 2) - math.pow(beta, 2)) / (2 * dist3d * gamma))
    # f_diff = (math.pow(dist3d, 2) + math.pow(gamma, 2) - math.pow(beta, 2)) / (2 * dist3d * gamma)
    f_diff = f_diff / math.pi * 180
    return case, azi_diff, tilt_diff, dist3d, azi_diff*dist3d


def lams(option, id):
    img_width = 40
    img_height = 40
    distance = 40
    try:
        os.stat("lams")
    except:
        os.mkdir("lams")
    try:
        os.stat("lams/" + id)
    except:
        os.mkdir("lams/" + id)
    grid_path = "lams/" + id
    excel_file = id + '.xlsx'
    excel = pd.read_excel(excel_file, engine='openpyxl')
    for i in range(len(excel)):
        Tx = [int(excel['TX'][i]), int(excel['TY'][i]), int(excel['TZ'][i])]
        Rx = [int(excel['RX'][i]), int(excel['RY'][i]), 2]
        chai = int(excel['chai1'][i])
        gen_image(Tx, Rx, grid_path + '/' + str(excel['idx'][i]), img_width, img_height, distance, ver=2, chai=chai)
        print(i + 1, ' / ', len(excel), ': GRID')


def predict(excel, lams_path):
    numeric_data_rss = pd.read_excel(excel, engine='openpyxl')
    model_name = '2928_5.8494.h5'
    model = load_model(model_name)
    length = len(numeric_data_rss)
    x_test = np.zeros((length, 4))

    test_lams_data = []
    lams_shape = [20, 20, 20]
    lams_data = []
    y=0
    for i in range(len(numeric_data_rss)):
        test_lams_data.append(np.load(lams_path + '/' + str(i + 1) + '.npy'))
        # test_lams_2ddata.append(np.load(lams_2dpath + '/' + str(i + 1) + '.npy'))
        x_test[y][0] = (180 - numeric_data_rss['chai1'][i]) / 180 * 20
        x_test[y][1] = (90 - numeric_data_rss['chai2'][i]) / 90
        x_test[y][2] = (262 - numeric_data_rss['dist'][i]) / 262
        x_test[y][3] = (17729 - numeric_data_rss['chaidist'][i]) / 17229
        y+=1

    numeric_test = np.array(x_test).astype(np.float32)
    img_test = np.array(test_lams_data).astype(np.float32)

    img_test = np.expand_dims(img_test, axis=4)
    predictions = model.predict([img_test, numeric_test])
    return predictions
    # evaluate = model.evaluate([img_test, numeric_test], label_test, verbose=2)


def visualization(grid_excel_path, result):
    image_name = './imagedata/allMap.png'
    temp = pd.read_excel(grid_excel_path, engine='openpyxl')
    RX = temp['RX']
    RX -= x_min
    RY = temp['RY']
    RY -= y_min
    RY = 3060 - RY
    Color = []
    for i in range(len(result)):
        Color.append(result[i][0])
    Color_min = result.min()
    Color_max = result.max()
    Color = Color - Color_min
    Color /= Color_max
    Color *= 255
    b = dim_1 - 255
    b *= -1
    index = []
    for i in range(len(Color)):
        index.append(i)
    #print(Color)
    Color = pd.Series(Color, index=index)
    im = plt.imread(image_name)
    implot = plt.imshow(im, cmap='gray')

    plt.scatter(x=RX, y=RY, c=Color, s=10, cmap='coolwarm')
    fig, ax = plt.subplots()
    #plt.figure(figsize=(50, 50))
    #plt.savefig('Results.png')
    #plt.colorbar()
    #plt.show()
    plt.savefig('./resultImage/Result.png')


def main():
    dim_1, dim_2 = load_npy()
    cnt = 1
    write_wb = Workbook()
    write_ws = write_wb.active
    write_ws['A1'] = 'idx'
    write_ws['B1'] = 'RX'
    write_ws['C1'] = 'RY'
    write_ws['D1'] = 'RZ'
    write_ws['E1'] = 'TX'
    write_ws['F1'] = 'TY'
    write_ws['G1'] = 'TZ'
    write_ws['H1'] = 'chai1'
    write_ws['I1'] = 'chai2'
    write_ws['J1'] = 'dist'
    write_ws['K1'] = 'chaidist'
    id = sys.argv[1]
    TX = sys.argv[2]
    TY = sys.argv[3]
    azimuth = sys.argv[4]
    downtilt = sys.argv[5]

    idx = 2

    for y in range(4151010, 4151310, 20):
        for x in range(325914, 326290, 20):
            if dim_1[y - y_min][x - x_min] == 0:
                write_ws.cell(idx, 1, idx-1)
                write_ws.cell(idx, 2, x)
                write_ws.cell(idx, 3, y)
                write_ws.cell(idx, 4, dim_1[y - y_min][x - x_min]+2)
                write_ws.cell(idx, 5, TX)
                write_ws.cell(idx, 6, TY)
                write_ws.cell(idx, 7, dim_2[int(TY)-y_min][int(TX)-x_min]+5)
                try:
                    Tx = [int(TX), int(TY), dim_2[int(TY)-y_min][int(TX)-x_min]+5]
                except:
                    print("건물위가 아닙니다")
                    exit()
                azimuth = int(azimuth)
                downtilt = int(downtilt)
                Rx = [x, y, dim_1[y-y_min][x-x_min]+2]
                res = angle_diff(Tx, Rx, azimuth, downtilt)
                # print(res)
                write_ws.cell(idx, 8, res[1])
                write_ws.cell(idx, 9, res[2])
                write_ws.cell(idx, 10, res[3])
                write_ws.cell(idx, 11, res[4])
                idx += 1

    # Grid 좌표 엑셀 파일 이름 설정해줘야함
    write_wb.save(id + '.xlsx')
    lams(0, id)
    res = predict(id + '.xlsx', "lams/" + id)
    visualization(id + '.xlsx', res)
    #print(res)


if __name__ == "__main__":
    main()
